# Import io for memory byte buffer operations
import io
# Import datetime for date formatting and arithmetic
from datetime import datetime
# Import relativedelta for month-based date arithmetic
from dateutil.relativedelta import relativedelta
# Import openpyxl library for Excel file creation and styling
import openpyxl
# Import openpyxl formatting components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# Import helper function to get column letter from number
from openpyxl.utils import get_column_letter
# Import pandas library for tabular data structures
import pandas as pd
# Import streamlit web application framework
import streamlit as st
# Import Supabase database client
from supabase import Client, create_client

# ==========================================
# 1. DATABASE & APP INITIALIZATION
# ==========================================
# Retrieve Supabase project URL from Streamlit secrets
SUPABASE_URL = st.secrets["SUPABASE_URL"]
# Retrieve Supabase public anon key from Streamlit secrets
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
# Instantiate connection client to Supabase database
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Retrieve Committee PIN from secrets (fallback default 8899)
COMMITTEE_PIN = str(st.secrets.get("COMMITTEE_PIN", "8899"))

# Initialize session state for committee login
if "committee_authenticated" not in st.session_state:
    # Set initial authentication status to False
    st.session_state["committee_authenticated"] = False

# Set page configuration for wide screen layout
st.set_page_config(page_title="Sunway Sinar Maintenance Portal", layout="wide")

# Define standard monthly maintenance fee amount
MONTHLY_FEE = 35.00
# Define annual total expected fee (12 months * RM 35.00 = RM 420.00)
ANNUAL_EXPECTED_FEE = 12 * MONTHLY_FEE

# Standard Month Names List
MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
# Multi-year selection range (supporting 2019 onwards)
YEAR_OPTIONS = list(range(2019, 2036))
# Get active calendar year dynamically
CURRENT_YEAR = datetime.now().year

# Natural, friendly everyday language dictionary
TRANSLATIONS = {
    "en": {
        "app_title": "🏢 Sunway Sinar Maintenance Fee Portal",
        "tab1_title": "🔍 Payment Status",
        "tab2_title": "📝 Record Payment",
        "tab3_title": "📅 Monthly Collection",
        "tab4_title": "📊 Annual Summary",
        "lock_warning": "🔒 This section is restricted to resident committee members.",
        "pin_placeholder": "Enter Committee PIN to Unlock",
        "unlock_btn": "🔓 Unlock Access",
        "lock_logout_btn": "🔒 Lock / Log Out",
        "pin_success": "✅ PIN Verified! Unlocking access...",
        "pin_error": "❌ Incorrect PIN. Please contact the management committee.",
        
        # Tab 1
        "tab1_header": "🔍 Check Unit Payment Status",
        "tab1_caption": "Select your unit to check monthly payment records, outstanding balance, and official receipt history.",
        "select_block": "Select Block",
        "select_floor": "Select Floor Level",
        "select_unit": "Select Unit Number",
        "view_year": "View Year",
        "ground_floor": "Ground Floor",
        "level": "Level",
        "unit_status_title": "Unit Status",
        "annual_fee_req": "Annual Fee Required",
        "total_paid": "Total Paid",
        "outstanding_bal": "Outstanding Balance",
        "fully_settled": "Fully Settled (No Due)",
        "unpaid": "Unpaid",
        "history_title": "🧾 Payment Receipt History",
        "no_history": "No payment records found for this unit.",
        "col_pay_date": "Payment Date",
        "col_or_no": "Official Receipt (OR NO.)",
        "col_method": "Payment Method",
        "col_amount": "Amount (RM)",
        "col_start": "Start Month",
        "col_end": "End Month",
        "col_collector": "Collector Name",
        
        # Tab 2
        "tab2_header": "📝 Record Maintenance Payment",
        "collector_name_label": "Collector Name",
        "collector_default": "Committee Member",
        "new_entry_for": "New Payment Entry for Unit",
        "or_no_placeholder": "e.g. 0006",
        "for_year": "For Year",
        "unpaid_month_label": "Available Unpaid Month",
        "all_paid_msg": "🎉 This unit has fully settled all maintenance fees for this year!",
        "pay_method_label": "Payment Method",
        "amount_received_label": "Amount Received (RM)",
        "submit_btn": "Submit Payment Record",
        "payment_success": "Payment of RM {amt:.2f} recorded for {month} {year} on unit {unit}!",
        "recent_entries_title": "🛠️ Search & Manage Payment Records",
        "save_changes_btn": "💾 Save Changes",
        "delete_record_btn": "🗑️ Delete Record",
        "search_mgmt_label": "🔍 Search Receipt / Unit / Collector",
        "search_mgmt_placeholder": "e.g. 0006, S2-216, Nana",
        "show_label": "Show",
        "sort_order_label": "Sort Order",
        "opt_desc": "Newest First (Latest)",
        "opt_asc": "Oldest First",
        "showing_mgmt_caption": "Showing **{shown}** of **{total}** matching records.",
        "no_mgmt_found": "No payment records found matching your search.",
        
        # Tab 3
        "tab3_header": "📅 Monthly Collection Records",
        "select_month": "Select Month",
        "select_year": "Select Year",
        "sort_by": "Sort By",
        "order_by": "Sort Order",
        "sort_opt_date": "Payment Date",
        "sort_opt_receipt": "Receipt Number (OR NO.)",
        "sort_opt_unit": "Unit ID",
        "sort_opt_block": "Block",
        "sort_opt_floor": "Floor Level",
        "sort_opt_out": "Outstanding Balance",
        "sort_opt_total": "Total Paid",
        "order_asc": "Ascending (Oldest / Lowest / A-Z)",
        "order_desc": "Descending (Newest / Highest / Z-A)",
        "no_monthly_records": "No payment records found for {month} {year}.",
        "found_logs": "Found **{count}** payment records.",
        "dl_monthly_btn": "📥 Download {month} {year} Collection Record (.xlsx)",
        "col_logged_date": "Recorded Date",
        
        # Tab 4
        "tab4_header": "📊 Annual Payment Summary & Outstanding Tracker",
        "filter_block": "Filter Block",
        "all_blocks": "All Blocks",
        "search_unit": "🔍 Search Unit",
        "search_placeholder": "e.g. 216, G01, S2",
        "showing_units_msg": "Showing **{count}** units matching filter/search for year **{year}**.",
        "dl_annual_btn": "📥 Download Summary for {block} ({year}) (.xlsx)"
    },
    "ms": {
        "app_title": "🏢 Portal Yuran Penyelenggaraan Sunway Sinar",
        "tab1_title": "🔍 Semakan Bayaran",
        "tab2_title": "📝 Rekod Bayaran",
        "tab3_title": "📅 Kutipan Bulanan",
        "tab4_title": "📊 Ringkasan Tahunan",
        "lock_warning": "🔒 Bahagian ini dikhaskan untuk Ahli Jawatankuasa (AJK) sahaja.",
        "pin_placeholder": "Masukkan PIN AJK untuk Akses",
        "unlock_btn": "🔓 Buka Kunci Akses",
        "lock_logout_btn": "🔒 Kunci / Log Keluar",
        "pin_success": "✅ PIN Sah! Membuka akses...",
        "pin_error": "❌ PIN tidak sah. Sila hubungi pihak pengurusan.",
        
        # Tab 1
        "tab1_header": "🔍 Semak Status Bayaran Unit",
        "tab1_caption": "Sila pilih unit anda untuk menyemak rekod bayaran bulanan, baki tunggakan, dan senarai resit bayaran.",
        "select_block": "Pilih Blok",
        "select_floor": "Pilih Aras / Tingkat",
        "select_unit": "Pilih Nombor Unit",
        "view_year": "Pilih Tahun",
        "ground_floor": "Aras Bawah (Ground)",
        "level": "Aras",
        "unit_status_title": "Status Unit",
        "annual_fee_req": "Jumlah Yuran Setahun",
        "total_paid": "Jumlah Sudah Dibayar",
        "outstanding_bal": "Baki Tunggakan",
        "fully_settled": "Selesai (Tiada Tunggakan)",
        "unpaid": "Belum Bayar",
        "history_title": "🧾 Sejarah Resit Bayaran",
        "no_history": "Tiada rekod bayaran dijumpai untuk unit ini.",
        "col_pay_date": "Tarikh Bayar",
        "col_or_no": "No. Resit (OR NO.)",
        "col_method": "Kaedah Bayaran",
        "col_amount": "Jumlah (RM)",
        "col_start": "Bulan Mula",
        "col_end": "Bulan Akhir",
        "col_collector": "Nama Pengutip",
        
        # Tab 2
        "tab2_header": "📝 Rekod Bayaran Penyelenggaraan",
        "collector_name_label": "Nama Pengutip / AJK",
        "collector_default": "Ahli Jawatankuasa",
        "new_entry_for": "Rekod Bayaran Baru Bagi Unit",
        "or_no_placeholder": "cth. 0006",
        "for_year": "Untuk Tahun",
        "unpaid_month_label": "Pilih Bulan Tertunggak",
        "all_paid_msg": "🎉 Unit ini telah selesai membayar semua yuran bagi tahun ini!",
        "pay_method_label": "Kaedah Bayaran",
        "amount_received_label": "Jumlah Diterima (RM)",
        "submit_btn": "Simpan Rekod Bayaran",
        "payment_success": "Bayaran RM {amt:.2f} berjaya disimpan bagi {month} {year} untuk unit {unit}!",
        "recent_entries_title": "🛠️ Carian & Urus Rekod Bayaran",
        "save_changes_btn": "💾 Simpan Perubahan",
        "delete_record_btn": "🗑️ Padam Rekod",
        "search_mgmt_label": "🔍 Cari No. Resit / Unit / Pengutip",
        "search_mgmt_placeholder": "cth. 0006, S2-216, Nana",
        "show_label": "Papar",
        "sort_order_label": "Susunan",
        "opt_desc": "Paling Baru Dahulu (Baru ➔ Lama)",
        "opt_asc": "Paling Lama Dahulu (Lama ➔ Baru)",
        "showing_mgmt_caption": "Memaparkan **{shown}** daripada **{total}** rekod dijumpai.",
        "no_mgmt_found": "Tiada rekod bayaran dijumpai sepadan dengan carian anda.",
        
        # Tab 3
        "tab3_header": "📅 Rekod Kutipan Bulanan",
        "select_month": "Pilih Bulan",
        "select_year": "Pilih Tahun",
        "sort_by": "Susun Mengikut",
        "order_by": "Susunan",
        "sort_opt_date": "Tarikh Bayar",
        "sort_opt_receipt": "No. Resit (OR NO.)",
        "sort_opt_unit": "No. Unit",
        "sort_opt_block": "Blok",
        "sort_opt_floor": "Aras Tingkat",
        "sort_opt_out": "Baki Tunggakan",
        "sort_opt_total": "Jumlah Bayaran",
        "order_asc": "Menaik (Lama ➔ Baru / A ➔ Z)",
        "order_desc": "Menurun (Baru ➔ Lama / Z ➔ A)",
        "no_monthly_records": "Tiada rekod kutipan pada bulan {month} {year}.",
        "found_logs": "Dijumpai sebanyak **{count}** rekod kutipan.",
        "dl_monthly_btn": "📥 Muat Turun Rekod Kutipan {month} {year} (.xlsx)",
        "col_logged_date": "Tarikh Direkod",
        
        # Tab 4
        "tab4_header": "📊 Ringkasan Tahunan & Semakan Tunggakan",
        "filter_block": "Tapis Mengikut Blok",
        "all_blocks": "Semua Blok",
        "search_unit": "🔍 Cari Unit",
        "search_placeholder": "cth. 216, G01, S2",
        "showing_units_msg": "Memaparkan **{count}** unit bagi tahun **{year}**.",
        "dl_annual_btn": "📥 Muat Turun Ringkasan {block} ({year}) (.xlsx)"
    }
}

# Top Header Layout: Title on Left, Compact Language Toggle on Right
head_col1, head_col2 = st.columns([4.2, 1.3])
with head_col2:
    # Select language choice
    lang_choice = st.selectbox("🌐 Language / Bahasa", ["English", "Bahasa Melayu"], index=0, label_visibility="collapsed")
    # Set active language code
    lang = "en" if lang_choice == "English" else "ms"
    # Assign active translation mapping
    t = TRANSLATIONS[lang]

with head_col1:
    # Render main header title
    st.title(t["app_title"])

# ==========================================
# 2. HELPER CALCULATION & EXPORT FUNCTIONS
# ==========================================
def get_unit_payments(unit_id):
    # Fetch all payment records belonging to a specific unit
    response = supabase.table("payments").select("*").eq("unit_id", unit_id).order("created_at", desc=True).execute()
    return response.data

def delete_payment_entry(rec):
    # Helper function to delete an incorrect payment record
    query = supabase.table("payments").delete()
    if "id" in rec and rec["id"] is not None:
        query = query.eq("id", rec["id"])
    else:
        query = query.eq("unit_id", rec["unit_id"]).eq("created_at", rec["created_at"])
    query.execute()
    st.rerun()

def update_payment_entry(rec, new_or, new_collector, new_method, new_amount):
    # Helper function to update existing payment record details
    payload = {
        "or_no": new_or,
        "collector_name": new_collector,
        "payment_method": new_method,
        "amount_paid": new_amount
    }
    query = supabase.table("payments").update(payload)
    if "id" in rec and rec["id"] is not None:
        query = query.eq("id", rec["id"])
    else:
        query = query.eq("unit_id", rec["unit_id"]).eq("created_at", rec["created_at"])
    query.execute()
    st.rerun()

def calculate_monthly_balances(records, target_year):
    # If no records exist, return zeros for all 12 months
    if not records:
        return {f"{target_year}-{m:02d}": 0.0 for m in range(1, 13)}

    # Sort payment records chronologically by creation date
    sorted_recs = sorted(records, key=lambda x: (x.get("created_at", ""), x.get("start_month", "")))
    
    # Pool all raw monthly payments across the entire timeline
    monthly_pool = {}
    for rec in sorted_recs:
        curr = datetime.strptime(rec["start_month"], "%Y-%m-%d")
        end = datetime.strptime(rec["end_month"], "%Y-%m-%d")
        amt = float(rec["amount_paid"])
        
        span_months = (end.year - curr.year) * 12 + (end.month - curr.month) + 1
        allocated = amt / max(span_months, 1)
        
        while curr <= end:
            m_key = curr.strftime("%Y-%m")
            monthly_pool[m_key] = monthly_pool.get(m_key, 0.0) + allocated
            curr += relativedelta(months=1)
            
    # Sequential Waterfall allocation across chronological months
    sorted_all_months = sorted(list(monthly_pool.keys()))
    start_dt = datetime.strptime(sorted_all_months[0], "%Y-%m")
    min_year = min(start_dt.year, target_year)
    max_year = max(datetime.strptime(sorted_all_months[-1], "%Y-%m").year, target_year)
    
    cur_dt = datetime(min_year, 1, 1)
    end_dt = datetime(max_year, 12, 1)
    
    waterfall_balances = {}
    carry_forward = 0.0
    
    # Traverse every month sequentially and roll forward surplus
    while cur_dt <= end_dt:
        m_key = cur_dt.strftime("%Y-%m")
        direct_amt = monthly_pool.get(m_key, 0.0)
        total_available = direct_amt + carry_forward
        
        if total_available >= MONTHLY_FEE:
            waterfall_balances[m_key] = MONTHLY_FEE
            carry_forward = total_available - MONTHLY_FEE
        else:
            waterfall_balances[m_key] = total_available
            carry_forward = 0.0
            
        cur_dt += relativedelta(months=1)
        
    # Extract calculated values for the target year
    return {f"{target_year}-{m:02d}": waterfall_balances.get(f"{target_year}-{m:02d}", 0.0) for m in range(1, 13)}

def generate_monthly_excel(records, month_name, year):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year} Log"

    # Merge title banner across columns A to I
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"SUNWAY SINAR MAINTENANCE FEE COLLECTION - {month_name.upper()} {year}"
    title_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["No", "Collection Date", "OR NO.", "Unit ID", "Collector Name", "Payment Method", "Amount Paid (RM)", "Coverage Start", "Coverage End"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.row_dimensions[3].height = 22
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, rec in enumerate(records, start=1):
        row_num = idx + 3
        amt = float(rec["amount_paid"])
        row_data = [
            idx,
            rec["created_at"][:10],
            rec.get("or_no", "-"),
            rec["unit_id"],
            rec.get("collector_name", "-"),
            rec.get("payment_method", "CASH"),
            amt,
            datetime.strptime(rec["start_month"], "%Y-%m-%d").strftime("%b %Y"),
            datetime.strptime(rec["end_month"], "%Y-%m-%d").strftime("%b %Y")
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20
        
        for c_idx in range(1, 10):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            if c_idx == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    if len(records) > 0:
        tot_row = len(records) + 4
        ws.cell(row=tot_row, column=6, value="TOTAL COLLECTED:").font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=tot_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
        tot_cell = ws.cell(row=tot_row, column=7, value=f"=SUM(G4:G{tot_row-1})")
        tot_cell.font = Font(name="Arial", size=10, bold=True)
        tot_cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_annual_excel(matrix_df, year):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Summary {year}"

    ws.merge_cells("A1:O1")
    t_cell = ws["A1"]
    t_cell.value = f"SUNWAY SINAR MAINTENANCE FEE - ANNUAL SUMMARY ({year})"
    t_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = list(matrix_df.columns)
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    paid_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    partial_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    unpaid_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    paid_font = Font(name="Arial", size=9, bold=True, color="375623")
    partial_font = Font(name="Arial", size=9, bold=True, color="B25900")
    unpaid_font = Font(name="Arial", size=9, bold=True, color="C65911")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.row_dimensions[3].height = 22
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in matrix_df.iterrows():
        r_num = idx + 4
        ws.append(list(row))
        ws.row_dimensions[r_num].height = 20
        
        ws.cell(row=r_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_num, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r_num, column=1).border = thin_border

        for c_idx in range(2, 14):
            val = row.iloc[c_idx - 1]
            cell = ws.cell(row=r_num, column=c_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "PAID":
                cell.fill = paid_fill
                cell.font = paid_font
            elif "PARTIAL" in str(val):
                cell.fill = partial_fill
                cell.font = partial_font
            else:
                cell.fill = unpaid_fill
                cell.font = unpaid_font
                
        for c_idx in [14, 15]:
            cell = ws.cell(row=r_num, column=c_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.number_format = '#,##0.00'

    if len(matrix_df) > 0:
        tot_row = len(matrix_df) + 4
        ws.cell(row=tot_row, column=1, value="TOTAL:").font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        tot_paid_cell = ws.cell(row=tot_row, column=14, value=f"=SUM(N4:N{tot_row-1})")
        tot_paid_cell.font = Font(name="Arial", size=10, bold=True)
        tot_paid_cell.number_format = '#,##0.00'
        
        tot_out_cell = ws.cell(row=tot_row, column=15, value=f"=SUM(O4:O{tot_row-1})")
        tot_out_cell.font = Font(name="Arial", size=10, bold=True)
        tot_out_cell.number_format = '#,##0.00'

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 13

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def get_unpaid_months_for_year(unit_id, target_year):
    records = get_unit_payments(unit_id)
    balances = calculate_monthly_balances(records, target_year)
    available_months = []
    for m in range(1, 13):
        m_key = f"{target_year}-{m:02d}"
        if balances.get(m_key, 0.0) < MONTHLY_FEE:
            available_months.append(MONTH_NAMES[m - 1])
    return available_months

# Helper function to gate committee access using PIN authentication
def verify_committee_access(tab_name_key):
    # Check if committee session state is already authenticated
    if st.session_state.get("committee_authenticated", False):
        # Position logout button in top-right corner
        col_status, col_btn = st.columns([5, 1.2])
        with col_btn:
            # Button to lock all tabs simultaneously
            if st.button("🔒 " + ("Kunci Semula" if lang == "ms" else "Lock / Log Out"), key=f"btn_logout_{tab_name_key}", use_container_width=True):
                # Reset global session state
                st.session_state["committee_authenticated"] = False
                # Refresh page to apply lock across all tabs
                st.rerun()
        return True

    # Centered PIN card layout
    spacer_left, center_col, spacer_right = st.columns([1, 1.2, 1])
    with center_col:
        # Display restriction warning notice inside the card
        st.warning(t["lock_warning"])
        
        # Centered password input field
        pin_input = st.text_input(
            t["pin_placeholder"], 
            type="password", 
            key=f"auth_pin_input_{tab_name_key}_{lang}"
        )
        
        # Centered unlock button placed directly underneath the input box
        if st.button(t["unlock_btn"], type="primary", key=f"btn_unlock_pin_{tab_name_key}_{lang}", use_container_width=True):
            # Validate input PIN against secret configuration
            if pin_input.strip() == COMMITTEE_PIN.strip():
                # Shared global state: Unlocks Tabs 2, 3, and 4 simultaneously
                st.session_state["committee_authenticated"] = True
                # Display success message
                st.success(t["pin_success"])
                # Rerun application to render unlocked content
                st.rerun()
            else:
                # Display error alert on invalid PIN entry
                st.error(t["pin_error"])
                
    return False

# ==========================================
# 3. FOUR INTERFACE TABS
# ==========================================
tab_status, tab_entry, tab_monthly, tab_annual = st.tabs([
    t["tab1_title"],
    t["tab2_title"],
    t["tab3_title"],
    t["tab4_title"]
])

# ==========================================
# --- TAB 1: 🔍 PAYMENT STATUS (RESIDENT / INQUIRY) ---
# ==========================================
with tab_status:
    st.header(t["tab1_header"])
    st.caption(t["tab1_caption"])
    
    # Unit selection controls
    u_col1, u_col2, u_col3, u_col4 = st.columns(4)
    with u_col1:
        s_block = st.selectbox(t["select_block"], ["S1", "S2", "S3"], key="status_block")
    with u_col2:
        s_floor = st.selectbox(
            t["select_floor"], 
            [0, 1, 2, 3, 4], 
            format_func=lambda x: t["ground_floor"] if x == 0 else f"{t['level']} {x}", 
            key="status_floor"
        )
    with u_col3:
        units_resp = supabase.table("units").select("unit_id").eq("block", s_block).eq("floor_level", s_floor).execute()
        unit_list = [u["unit_id"] for u in units_resp.data] if units_resp.data else []
        s_unit = st.selectbox(t["select_unit"], unit_list, key="status_unit") if unit_list else None
    with u_col4:
        s_view_year = st.selectbox(t["view_year"], YEAR_OPTIONS, index=YEAR_OPTIONS.index(CURRENT_YEAR), key="status_year")

    st.markdown("---")
    
    if s_unit:
        unit_records = get_unit_payments(s_unit)
        monthly_balances = calculate_monthly_balances(unit_records, s_view_year)
        
        total_paid_unit = sum(monthly_balances.values())
        total_outstanding_unit = max(0.0, ANNUAL_EXPECTED_FEE - total_paid_unit)
        
        st.subheader(f"{t['unit_status_title']}: {s_unit} ({s_view_year})")
        
        # Metric cards summary
        m_card1, m_card2, m_card3 = st.columns(3)
        m_card1.metric(f"{t['annual_fee_req']} ({s_view_year})", f"RM {ANNUAL_EXPECTED_FEE:.2f}")
        m_card2.metric(f"{t['total_paid']} ({s_view_year})", f"RM {total_paid_unit:.2f}")
        m_card3.metric(
            f"{t['outstanding_bal']} ({s_view_year})", 
            f"RM {total_outstanding_unit:.2f}", 
            delta=f"-RM {total_outstanding_unit:.2f}" if total_outstanding_unit > 0 else t["fully_settled"], 
            delta_color="inverse"
        )
        
        st.write("")
        # Display monthly status indicators
        view_year_months = [datetime(s_view_year, m, 1) for m in range(1, 13)]
        
        # Row 1: January to June
        r1_cols = st.columns(6)
        for idx in range(6):
            m_date = view_year_months[idx]
            m_str = m_date.strftime("%Y-%m")
            m_lbl = m_date.strftime('%b')
            paid_amount = monthly_balances.get(m_str, 0.0)
            
            if paid_amount >= MONTHLY_FEE:
                r1_cols[idx].success(f"✅ **{m_lbl}**: RM{paid_amount:.2f}")
            elif paid_amount > 0:
                r1_cols[idx].warning(f"⚠️ **{m_lbl}**: RM{paid_amount:.2f}/{MONTHLY_FEE:.0f}")
            else:
                r1_cols[idx].error(f"❌ **{m_lbl}**: {t['unpaid']}")

        # Row 2: July to December
        r2_cols = st.columns(6)
        for idx in range(6, 12):
            m_date = view_year_months[idx]
            m_str = m_date.strftime("%Y-%m")
            m_lbl = m_date.strftime('%b')
            paid_amount = monthly_balances.get(m_str, 0.0)
            
            if paid_amount >= MONTHLY_FEE:
                r2_cols[idx - 6].success(f"✅ **{m_lbl}**: RM{paid_amount:.2f}")
            elif paid_amount > 0:
                r2_cols[idx - 6].warning(f"⚠️ **{m_lbl}**: RM{paid_amount:.2f}/{MONTHLY_FEE:.0f}")
            else:
                r2_cols[idx - 6].error(f"❌ **{m_lbl}**: {t['unpaid']}")

        st.markdown("---")
        # Personal Receipt History for this Unit
        st.subheader(f"{t['history_title']} ({s_unit})")
        if len(unit_records) == 0:
            st.info(t["no_history"])
        else:
            df_unit_history = pd.DataFrame(unit_records)[["created_at", "or_no", "payment_method", "amount_paid", "start_month", "end_month", "collector_name"]]
            df_unit_history.columns = [
                t["col_pay_date"], t["col_or_no"], t["col_method"], 
                t["col_amount"], t["col_start"], t["col_end"], t["col_collector"]
            ]
            df_unit_history[t["col_pay_date"]] = df_unit_history[t["col_pay_date"]].apply(lambda x: str(x)[:10])
            df_unit_history[t["col_amount"]] = df_unit_history[t["col_amount"]].apply(lambda x: f"RM {float(x):.2f}")
            st.dataframe(df_unit_history, use_container_width=True)

# ==========================================
# --- TAB 2: 📝 RECORD PAYMENT (COMMITTEE ENTRY) ---
# ==========================================
with tab_entry:
    st.header(t["tab2_header"])
    if verify_committee_access("tab2"):
        # Committee collector and unit selection
        e_col1, e_col2, e_col3, e_col4 = st.columns(4)
        with e_col1:
            e_block = st.selectbox(t["select_block"], ["S1", "S2", "S3"], key="entry_block")
        with e_col2:
            e_floor = st.selectbox(
                t["select_floor"], 
                [0, 1, 2, 3, 4], 
                format_func=lambda x: t["ground_floor"] if x == 0 else f"{t['level']} {x}", 
                key="entry_floor"
            )
        with e_col3:
            units_resp_e = supabase.table("units").select("unit_id").eq("block", e_block).eq("floor_level", e_floor).execute()
            unit_list_e = [u["unit_id"] for u in units_resp_e.data] if units_resp_e.data else []
            selected_unit_e = st.selectbox(t["select_unit"], unit_list_e, key="entry_unit") if unit_list_e else None
        with e_col4:
            collector = st.text_input(t["collector_name_label"], value=t["collector_default"], key="entry_collector")

        st.markdown("---")
        if selected_unit_e:
            st.subheader(f"{t['new_entry_for']}: {selected_unit_e}")
            
            # Row 1: Receipt Number (Left) | Target Year & Available Unpaid Months (Right)
            r1_col1, r1_col2, r1_col3 = st.columns([2, 1, 1])
            with r1_col1:
                or_no = st.text_input(t["col_or_no"], placeholder=t["or_no_placeholder"], key=f"entry_or_no_{selected_unit_e}")
            with r1_col2:
                entry_year_val = st.selectbox(t["for_year"], YEAR_OPTIONS, index=YEAR_OPTIONS.index(CURRENT_YEAR), key=f"entry_yr_{selected_unit_e}")
            
            # Calculate only unpaid months for selected year
            unpaid_months = get_unpaid_months_for_year(selected_unit_e, entry_year_val)
            
            with r1_col3:
                if unpaid_months:
                    start_month_name = st.selectbox(t["unpaid_month_label"], unpaid_months, index=0, key=f"entry_sm_{selected_unit_e}_{entry_year_val}")
                else:
                    st.selectbox(t["unpaid_month_label"], [t["fully_settled"]], disabled=True, key=f"entry_sm_disabled_{selected_unit_e}")
                    start_month_name = None

            if not unpaid_months:
                st.success(t["all_paid_msg"])
            else:
                # Row 2: Payment Method (Left) | Amount Received (Right)
                r2_col1, r2_col2 = st.columns([2, 2])
                with r2_col1:
                    payment_method = st.selectbox(t["pay_method_label"], ["CASH", "Online Transfer", "DuitNow QR", "CHEQUE"], key=f"entry_method_{selected_unit_e}")
                with r2_col2:
                    amount = st.number_input(t["amount_received_label"], min_value=0.0, step=5.0, value=35.0, key=f"entry_amount_{selected_unit_e}")

                start_m_idx = MONTH_NAMES.index(start_month_name) + 1
                start_m = datetime(entry_year_val, start_m_idx, 1)

                if st.button(t["submit_btn"], type="primary"):
                    payment_payload = {
                        "unit_id": selected_unit_e,
                        "or_no": or_no,
                        "collector_name": collector,
                        "payment_method": payment_method,
                        "amount_paid": amount,
                        "start_month": start_m.strftime("%Y-%m-%d"),
                        "end_month": start_m.strftime("%Y-%m-%d")
                    }
                    supabase.table("payments").insert(payment_payload).execute()
                    st.success(t["payment_success"].format(amt=amount, month=start_month_name, year=entry_year_val, unit=selected_unit_e))
                    st.rerun()

        # Add horizontal divider
        st.markdown("---")
        # Multi-language section title
        st.subheader(t["recent_entries_title"])

        # 3-column control bar for filtering
        ctl_c1, ctl_c2, ctl_c3 = st.columns([2, 1.2, 1.4])
        
        # 1. Search keyword input
        with ctl_c1:
            search_kw = st.text_input(
                t["search_mgmt_label"],
                placeholder=t["search_mgmt_placeholder"],
                key=f"manage_search_kw_{lang}"
            )
            
        # 2. Row limit dropdown
        with ctl_c2:
            limit_options = ["5", "10", "20", "50", "Semua" if lang == "ms" else "All"]
            limit_choice = st.selectbox(
                t["show_label"],
                limit_options,
                index=0,
                key=f"manage_limit_choice_{lang}"
            )
            
        # 3. Sorting order dropdown
        with ctl_c3:
            sort_opts = [t["opt_desc"], t["opt_asc"]]
            sort_choice = st.selectbox(
                t["sort_order_label"],
                sort_opts,
                index=0,
                key=f"manage_sort_choice_{lang}"
            )

        # Determine sorting direction from choice
        is_desc = True if sort_choice == t["opt_desc"] else False

        # Query Supabase payments table
        query_builder = supabase.table("payments").select("*").order("created_at", desc=is_desc)
        all_mgmt_data = query_builder.execute().data or []

        # Filter payment records based on search query
        if search_kw.strip():
            q_term = search_kw.strip().lower()
            filtered_records = [
                r for r in all_mgmt_data 
                if q_term in str(r.get("or_no", "")).lower() 
                or q_term in str(r.get("unit_id", "")).lower() 
                or q_term in str(r.get("collector_name", "")).lower()
            ]
        else:
            filtered_records = all_mgmt_data

        # Apply row limit
        if limit_choice not in ["Semua", "All"]:
            display_records = filtered_records[:int(limit_choice)]
        else:
            display_records = filtered_records

        # Summary record count
        st.caption(t["showing_mgmt_caption"].format(shown=len(display_records), total=len(filtered_records)))

        # Render interactive record cards
        if display_records:
            for idx, rec in enumerate(display_records):
                r_unit = rec.get("unit_id", "-")
                r_created_at = rec.get("created_at", "")
                r_date = str(r_created_at)[:16].replace("T", " ")
                r_or = rec.get("or_no", "-")
                r_amt = float(rec.get("amount_paid", 0.0))
                r_method = rec.get("payment_method", "CASH")
                r_collector = rec.get("collector_name", "-")
                
                card_label = f"📍 {r_unit} | Resit: {r_or} | RM {r_amt:.2f} ({r_date})" if lang == "ms" else f"📍 {r_unit} | Receipt: {r_or} | RM {r_amt:.2f} ({r_date})"
                with st.expander(card_label):
                    edit_c1, edit_c2, edit_c3, edit_c4 = st.columns(4)
                    with edit_c1:
                        new_or_val = st.text_input(t["col_or_no"], value=r_or, key=f"edit_or_{idx}_{r_created_at}_{lang}")
                    with edit_c2:
                        methods = ["CASH", "Online Transfer", "DuitNow QR", "CHEQUE"]
                        cur_idx = methods.index(r_method) if r_method in methods else 0
                        new_meth_val = st.selectbox(t["pay_method_label"], methods, index=cur_idx, key=f"edit_meth_{idx}_{r_created_at}_{lang}")
                    with edit_c3:
                        new_amt_val = st.number_input(t["col_amount"], value=r_amt, step=5.0, key=f"edit_amt_{idx}_{r_created_at}_{lang}")
                    with edit_c4:
                        new_col_val = st.text_input(t["col_collector"], value=r_collector, key=f"edit_col_{idx}_{r_created_at}_{lang}")
                    
                    btn_c1, btn_c2 = st.columns([1, 1])
                    with btn_c1:
                        if st.button(t["save_changes_btn"], key=f"btn_save_{idx}_{r_created_at}_{lang}", type="primary"):
                            update_payment_entry(rec, new_or_val, new_col_val, new_meth_val, new_amt_val)
                    with btn_c2:
                        if st.button(t["delete_record_btn"], key=f"btn_del_{idx}_{r_created_at}_{lang}"):
                            delete_payment_entry(rec)
        else:
            st.info(t["no_mgmt_found"])

# ==========================================
# --- TAB 3: 📅 MONTHLY COLLECTION EXPORT ---
# ==========================================
with tab_monthly:
    st.header(t["tab3_header"])
    if verify_committee_access("tab3"):
        m_col1, m_col2, m_col3, m_col4 = st.columns(4)
        with m_col1:
            selected_month_num = st.selectbox(
                t["select_month"], 
                list(range(1, 13)), 
                format_func=lambda x: datetime(CURRENT_YEAR, x, 1).strftime("%B") if lang == "en" else [
                    "Januari", "Februari", "Mac", "April", "Mei", "Jun", 
                    "Julai", "Ogos", "September", "Oktober", "November", "Disember"
                ][x - 1], 
                key="m_month"
            )
        with m_col2:
            selected_year = st.selectbox(t["select_year"], YEAR_OPTIONS, index=YEAR_OPTIONS.index(CURRENT_YEAR), key="m_year")
        with m_col3:
            sort_opts_m = [t["sort_opt_date"], t["sort_opt_receipt"], t["sort_opt_unit"], t["sort_opt_block"]]
            sort_by_m = st.selectbox(t["sort_by"], sort_opts_m, key="m_sort")
        with m_col4:
            order_opts_m = [t["order_asc"], t["order_desc"]]
            sort_order_m = st.selectbox(t["order_by"], order_opts_m, key="m_order")

        m_start_str = f"{selected_year}-{selected_month_num:02d}-01"
        next_m = datetime(selected_year, selected_month_num, 1) + relativedelta(months=1)
        m_end_str = next_m.strftime("%Y-%m-%d")

        records = supabase.table("payments").select("*").gte("created_at", m_start_str).lt("created_at", m_end_str).execute().data
        
        display_month_name = datetime(CURRENT_YEAR, selected_month_num, 1).strftime("%B") if lang == "en" else [
            "Januari", "Februari", "Mac", "April", "Mei", "Jun", 
            "Julai", "Ogos", "September", "Oktober", "November", "Disember"
        ][selected_month_num - 1]

        if len(records) == 0:
            st.info(t["no_monthly_records"].format(month=display_month_name, year=selected_year))
        else:
            st.write(t["found_logs"].format(count=len(records)))
            df_logs = pd.DataFrame(records)[["created_at", "or_no", "unit_id", "collector_name", "payment_method", "amount_paid", "start_month", "end_month"]]
            df_logs["Block"] = df_logs["unit_id"].apply(lambda x: x.split("-")[0] if "-" in str(x) else "")
            
            sort_map = {
                t["sort_opt_date"]: "created_at",
                t["sort_opt_receipt"]: "or_no",
                t["sort_opt_unit"]: "unit_id",
                t["sort_opt_block"]: "Block"
            }
            
            is_asc = True if sort_order_m == t["order_asc"] else False
            df_logs = df_logs.sort_values(by=sort_map[sort_by_m], ascending=is_asc)
            
            df_logs_display = df_logs.drop(columns=["Block"]).copy()
            df_logs_display.columns = [
                t["col_logged_date"], t["col_or_no"], "Unit ID", 
                t["col_collector"], t["col_method"], t["col_amount"], 
                t["col_start"], t["col_end"]
            ]
            
            df_logs_display[t["col_logged_date"]] = pd.to_datetime(df_logs_display[t["col_logged_date"]]).dt.strftime("%Y-%m-%d %H:%M")
            df_logs_display[t["col_amount"]] = df_logs_display[t["col_amount"]].apply(lambda x: f"RM {float(x):.2f}")
            
            st.dataframe(df_logs_display, use_container_width=True)
            
            sorted_records = df_logs.to_dict("records")
            excel_file = generate_monthly_excel(sorted_records, display_month_name, selected_year)
            
            st.download_button(
                label=t["dl_monthly_btn"].format(month=display_month_name, year=selected_year),
                data=excel_file,
                file_name=f"Collection_Log_{display_month_name}_{selected_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ==========================================
# --- TAB 4: 📊 ANNUAL PAYMENT SUMMARY ---
# ==========================================
with tab_annual:
    st.header(t["tab4_header"])
    if verify_committee_access("tab4"):
        f_col1, f_col2, f_col3, f_col4, f_col5 = st.columns([1.2, 1.2, 1.5, 1.5, 1.5])
        with f_col1:
            ann_year = st.selectbox(t["view_year"], YEAR_OPTIONS, index=YEAR_OPTIONS.index(CURRENT_YEAR), key="annual_year")
        with f_col2:
            block_options = [t["all_blocks"], "Block S1", "Block S2", "Block S3"]
            block_filter = st.selectbox(t["filter_block"], block_options, key="ann_block_filter")
        with f_col3:
            search_query = st.text_input(t["search_unit"], placeholder=t["search_placeholder"], key="ann_search")
        with f_col4:
            sort_opts_a = [t["sort_opt_unit"], t["sort_opt_floor"], t["sort_opt_out"], t["sort_opt_total"]]
            sort_by_a = st.selectbox(t["sort_by"], sort_opts_a, key="ann_sort")
        with f_col5:
            order_opts_a = [t["order_asc"], t["order_desc"]]
            sort_order_a = st.selectbox(t["order_by"], order_opts_a, key="ann_order")
        
        all_units = supabase.table("units").select("unit_id", "block", "floor_level").execute().data
        all_payments = supabase.table("payments").select("*").execute().data
        
        unit_payment_map = {}
        for p in all_payments:
            u = p["unit_id"]
            if u not in unit_payment_map:
                unit_payment_map[u] = []
            unit_payment_map[u].append(p)
                
        months_list = [datetime(ann_year, m, 1) for m in range(1, 13)]
        matrix_data = []
        
        for u_obj in all_units:
            u_id = u_obj["unit_id"]
            row = {
                "Unit ID": u_id,
                "Block": u_obj["block"],
                "Floor": u_obj["floor_level"]
            }
            
            u_records = unit_payment_map.get(u_id, [])
            u_balances = calculate_monthly_balances(u_records, ann_year)
            
            unit_total_paid = 0.0
            for m_date in months_list:
                m_str = m_date.strftime("%Y-%m")
                bal = u_balances.get(m_str, 0.0)
                unit_total_paid += bal
                
                if bal >= MONTHLY_FEE:
                    row[m_date.strftime("%b")] = "PAID"
                elif bal > 0.0:
                    row[m_date.strftime("%b")] = f"PARTIAL (RM{bal:.2f})"
                else:
                    row[m_date.strftime("%b")] = "UNPAID"
                    
            unit_outstanding = max(0.0, ANNUAL_EXPECTED_FEE - unit_total_paid)
            row["Total Paid (RM)"] = round(unit_total_paid, 2)
            row["Outstanding (RM)"] = round(unit_outstanding, 2)
            matrix_data.append(row)
            
        df_matrix = pd.DataFrame(matrix_data)
        
        if block_filter == "Block S1":
            df_matrix = df_matrix[df_matrix["Block"] == "S1"]
        elif block_filter == "Block S2":
            df_matrix = df_matrix[df_matrix["Block"] == "S2"]
        elif block_filter == "Block S3":
            df_matrix = df_matrix[df_matrix["Block"] == "S3"]

        if search_query.strip():
            q = search_query.strip().lower()
            df_matrix = df_matrix[df_matrix["Unit ID"].str.lower().str.contains(q)]

        is_asc_a = True if sort_order_a == t["order_asc"] else False
        if sort_by_a == t["sort_opt_floor"]:
            df_matrix = df_matrix.sort_values(by=["Floor", "Unit ID"], ascending=[is_asc_a, True])
        elif sort_by_a == t["sort_opt_out"]:
            df_matrix = df_matrix.sort_values(by=["Outstanding (RM)", "Unit ID"], ascending=[is_asc_a, True])
        elif sort_by_a == t["sort_opt_total"]:
            df_matrix = df_matrix.sort_values(by=["Total Paid (RM)", "Unit ID"], ascending=[is_asc_a, True])
        else:
            df_matrix = df_matrix.sort_values(by="Unit ID", ascending=is_asc_a)
            
        st.caption(t["showing_units_msg"].format(count=len(df_matrix), year=ann_year))
        df_display = df_matrix.drop(columns=["Block", "Floor"])
        st.dataframe(df_display, use_container_width=True, height=400)
        
        ann_excel = generate_annual_excel(df_display, ann_year)
        filter_label = block_filter.replace(" ", "_") if block_filter != t["all_blocks"] else "All_Blocks"
        st.download_button(
            label=t["dl_annual_btn"].format(block=block_filter, year=ann_year),
            data=ann_excel,
            file_name=f"Maintenance_Fee_Summary_{ann_year}_{filter_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
